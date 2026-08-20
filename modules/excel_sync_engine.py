# -*- coding: utf-8 -*-
"""
Module: excel_sync_engine.py
Mục đích: Động cơ Đồng bộ 2 Chiều Sổ cái Căn cứ Pháp lý Sống (Live Excel Ledger Sync).
Chức năng:
- Tự động chèn dòng văn bản mới vào sheet KHO_CAN_CU_MASTER.
- Kế thừa toàn bộ Tag nghiệp vụ từ văn bản cũ bị thay thế để không đứt gãy hồ sơ.
- Tự động tìm dòng văn bản cũ và chuyển trạng thái sang 🔴 Hết hiệu lực toàn bộ từ ngày...
- Tự động ghi vết biến động vào sheet CHANGELOG_AUDIT_LOG phục vụ thanh tra.
"""

import os
import openpyxl
from datetime import datetime
from typing import Dict, List, Any, Optional


class LegalExcelSyncEngine:
    """
    Động cơ tự động cập nhật Sổ cái Master Excel.
    """

    def __init__(self, excel_path: Optional[str] = None):
        self.excel_path = excel_path or os.path.join(
            os.path.dirname(__file__), "..", "Kho_Can_Cu_Phap_Ly.xlsx"
        )

    def sync_new_document(
        self,
        so_hieu: str,
        loai_vb: str,
        co_quan: str,
        ngay_bh: str,
        ngay_hl: str,
        linh_vuc: str,
        cau_can_cu: str,
        thay_the_cho: Optional[List[str]] = None,
        tags_bo_sung: Optional[List[str]] = None,
        thu_bac: int = 300,
        ghi_chu: str = ""
    ) -> Dict[str, Any]:
        """
        Thực hiện đồng bộ 1 văn bản mới vào Sổ cái Excel.
        """
        if not os.path.exists(self.excel_path):
            return {"success": False, "error": f"Không tìm thấy file {self.excel_path}"}

        wb = openpyxl.load_workbook(self.excel_path)
        ws_master = wb["KHO_CAN_CU_MASTER"] if "KHO_CAN_CU_MASTER" in wb.sheetnames else wb.active
        ws_log = wb["CHANGELOG_AUDIT_LOG"] if "CHANGELOG_AUDIT_LOG" in wb.sheetnames else None

        thay_the_list = thay_the_cho or []
        inherited_tags = set(tags_bo_sung or [])

        # 1. Quét tìm các văn bản cũ bị thay thế để đổi trạng thái
        replaced_doc_numbers = []
        for row in range(2, ws_master.max_row + 1):
            cell_so_hieu = str(ws_master.cell(row=row, column=2).value or "").strip()
            for old_doc in thay_the_list:
                if old_doc.lower() in cell_so_hieu.lower() or cell_so_hieu.lower() in old_doc.lower():
                    # Kế thừa tags
                    old_tags_val = str(ws_master.cell(row=row, column=9).value or "")
                    for t in old_tags_val.split(","):
                        if t.strip():
                            inherited_tags.add(t.strip())
                    
                    # Đổi trạng thái văn bản cũ
                    old_status = str(ws_master.cell(row=row, column=7).value or "")
                    new_status = f"🔴 Hết hiệu lực toàn bộ (bị thay thế bởi {so_hieu})"
                    ws_master.cell(row=row, column=7, value=new_status)
                    ws_master.cell(row=row, column=12, value=f"Bị thay thế bởi {so_hieu} từ ngày {ngay_hl}")
                    replaced_doc_numbers.append(cell_so_hieu)

                    # Ghi Log
                    if ws_log is not None:
                        now_str = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
                        ws_log.append([
                            now_str, "REPLACE_DOC", so_hieu, cell_so_hieu,
                            old_status, new_status, f"Ban hành {so_hieu}", "ReconWatchdog_AutoSync"
                        ])

        # 2. Kiểm tra xem văn bản mới đã tồn tại trong Sổ cái chưa
        doc_exists = False
        for row in range(2, ws_master.max_row + 1):
            cell_so_hieu = str(ws_master.cell(row=row, column=2).value or "").strip()
            if cell_so_hieu.lower() == so_hieu.lower():
                doc_exists = True
                break

        if not doc_exists:
            new_stt = ws_master.max_row
            tags_str = ",".join(sorted(inherited_tags)) if inherited_tags else "ALL"
            new_row = [
                new_stt, so_hieu, loai_vb, co_quan, ngay_bh, ngay_hl,
                "🟢 Còn hiệu lực", linh_vuc, tags_str, thu_bac, cau_can_cu, ghi_chu
            ]
            ws_master.append(new_row)

            if ws_log is not None:
                now_str = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
                ws_log.append([
                    now_str, "INSERT_NEW", so_hieu, "N/A",
                    "N/A", "🟢 Còn hiệu lực", f"Ban hành {so_hieu}", "ReconWatchdog_AutoSync"
                ])

        wb.save(self.excel_path)
        return {
            "success": True,
            "inserted_new": not doc_exists,
            "so_hieu": so_hieu,
            "replaced_docs": replaced_doc_numbers,
            "inherited_tags": list(inherited_tags)
        }
