# -*- coding: utf-8 -*-
"""
Unit test kiểm thử Module Đồng Bộ CSDL Excel (modules/legal_db_sync.py)
Kiểm tra chuẩn 14 cột, máy trạng thái vòng đời (cũ hết hiệu lực / mới đang hiệu lực) và chống trùng lặp.
"""

import unittest
import os
import sys
import openpyxl

# Thêm thư mục gốc vào path
sys.path.insert(0, os.path.abspath(os.path.join(os.path.dirname(__file__), "..")))

from modules.legal_db_sync import sync_legal_document_to_excel, HEADERS_14

TEST_EXCEL_PATH = os.path.join(os.path.dirname(__file__), "test_kho_phap_ly_temp.xlsx")


class TestLegalDBSync(unittest.TestCase):

    def setUp(self):
        if os.path.exists(TEST_EXCEL_PATH):
            os.remove(TEST_EXCEL_PATH)

    def tearDown(self):
        if os.path.exists(TEST_EXCEL_PATH):
            os.remove(TEST_EXCEL_PATH)

    def test_01_create_new_excel_with_14_headers(self):
        """Kiểm tra khởi tạo bảng Excel với đúng 14 tiêu đề chuẩn"""
        doc_data = {
            "so_hieu_clean": "63/2014/NĐ-CP",
            "title": "Nghị định quy định chi tiết thi hành Luật Đấu thầu",
            "trich_yeu": "Quy định chi tiết thi hành một số điều của Luật Đấu thầu",
            "co_quan": "Chính phủ",
            "ngay_ban_hanh": "26/06/2014",
            "ngay_hieu_luc": "15/08/2014",
            "linh_vuc": "Đấu thầu",
            "loai_van_ban": "Nghị định",
            "van_ban_thay_the": "",
            "chuyen_tiep_ngan": "Áp dụng cho các gói thầu phát hành trước năm 2024",
            "goi_thau_tags": ["ALL", "XD-01", "TV-04"],
            "telegraph_url": "https://telegra.ph/test-nd63"
        }
        res = sync_legal_document_to_excel(doc_data, excel_path=TEST_EXCEL_PATH)
        self.assertTrue(res)
        self.assertTrue(os.path.exists(TEST_EXCEL_PATH))

        wb = openpyxl.load_workbook(TEST_EXCEL_PATH)
        ws = wb.active
        self.assertEqual(ws.max_row, 2)  # 1 header + 1 data row
        self.assertEqual(ws.max_column, 14)

        headers = [ws.cell(row=1, column=c).value for c in range(1, 15)]
        self.assertEqual(headers, HEADERS_14)
        wb.close()

    def test_02_lifecycle_replacement_state_transition(self):
        """Kiểm tra máy trạng thái: Nạp NĐ 63 -> Nạp NĐ 24 thay thế NĐ 63 -> NĐ 63 đổi thành 'Hết hiệu lực'"""
        # 1. Nạp NĐ 63/2014
        doc_old = {
            "so_hieu_clean": "63/2014/NĐ-CP",
            "title": "Nghị định 63/2014/NĐ-CP",
            "co_quan": "Chính phủ",
            "ngay_ban_hanh": "26/06/2014",
            "ngay_hieu_luc": "15/08/2014",
            "linh_vuc": "Đấu thầu",
            "loai_van_ban": "Nghị định",
            "van_ban_thay_the": ""
        }
        sync_legal_document_to_excel(doc_old, excel_path=TEST_EXCEL_PATH)

        # 2. Nạp NĐ 24/2024 thay thế NĐ 63/2014
        doc_new = {
            "so_hieu_clean": "24/2024/NĐ-CP",
            "title": "Nghị định 24/2024/NĐ-CP",
            "co_quan": "Chính phủ",
            "ngay_ban_hanh": "27/02/2024",
            "ngay_hieu_luc": "27/02/2024",
            "linh_vuc": "Đấu thầu",
            "loai_van_ban": "Nghị định",
            "van_ban_thay_the": "Nghị định số 63/2014/NĐ-CP",
            "chuyen_tiep_ngan": "Gói thầu mở trước 27/02/2024 tiếp tục thực hiện theo NĐ 63",
            "goi_thau_tags": ["ALL", "XD-01"]
        }
        sync_legal_document_to_excel(doc_new, excel_path=TEST_EXCEL_PATH)

        # 3. Kiểm tra kết quả trong Excel
        wb = openpyxl.load_workbook(TEST_EXCEL_PATH)
        ws = wb.active
        self.assertEqual(ws.max_row, 3)  # 1 header + 2 văn bản

        # Dòng 2 (NĐ 63): Phải đổi thành "Hết hiệu lực"
        self.assertEqual(ws.cell(row=2, column=4).value, "63/2014/NĐ-CP")
        self.assertEqual(ws.cell(row=2, column=9).value, "Hết hiệu lực")
        self.assertIn("24/2024/NĐ-CP", str(ws.cell(row=2, column=10).value))

        # Dòng 3 (NĐ 24): Phải là "Đang có hiệu lực"
        self.assertEqual(ws.cell(row=3, column=4).value, "24/2024/NĐ-CP")
        self.assertEqual(ws.cell(row=3, column=9).value, "Đang có hiệu lực")
        wb.close()

    def test_03_idempotency_no_duplicates(self):
        """Kiểm tra chống trùng lặp: Nạp 2 lần cùng 1 số hiệu không sinh thêm dòng thừa"""
        doc = {
            "so_hieu_clean": "06/2024/TT-BKHĐT",
            "title": "Thông tư 06/2024/TT-BKHĐT",
            "co_quan": "Bộ Kế hoạch và Đầu tư",
            "ngay_ban_hanh": "26/04/2024",
            "ngay_hieu_luc": "26/04/2024",
            "linh_vuc": "Đấu thầu qua mạng",
            "loai_van_ban": "Thông tư"
        }
        sync_legal_document_to_excel(doc, excel_path=TEST_EXCEL_PATH)
        sync_legal_document_to_excel(doc, excel_path=TEST_EXCEL_PATH)

        wb = openpyxl.load_workbook(TEST_EXCEL_PATH)
        ws = wb.active
        self.assertEqual(ws.max_row, 2)  # 1 header + duy nhất 1 dòng
        wb.close()


if __name__ == "__main__":
    unittest.main()
