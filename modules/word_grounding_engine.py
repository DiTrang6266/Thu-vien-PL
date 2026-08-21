# -*- coding: utf-8 -*-
"""
Module: word_grounding_engine.py
Mục đích: Động cơ Lọc Căn cứ Pháp lý Sống & Ốp tự động vào Phôi Word (Word Legal Grounding Injector).
Chức năng:
- Đọc Sổ cái Master Excel Kho_Can_Cu_Phap_Ly.xlsx (Chuẩn hóa 14 cột).
- Lọc theo Tag nghiệp vụ và Gói thầu (TV-04, TV-05, XD-01, TO_TRINH_DU_TOAN, TO_TRINH_KHLCNT...).
- Chỉ lấy các văn bản 🟢 Còn hiệu lực, loại bỏ 100% văn bản 🔴 Hết hiệu lực.
- Tự động sắp xếp theo Thứ bậc Lập pháp chuẩn Luật Ban hành VBQPPL:
  Luật (100) -> Nghị định (200) -> Thông tư (300) -> Quyết định/Quy chuẩn (400) -> Quyết định CĐT (500) -> Căn cứ thực tế (999).
"""

import os
import sys
import openpyxl
from datetime import datetime
from typing import List, Dict, Any, Optional

if hasattr(sys.stdout, 'reconfigure'):
    sys.stdout.reconfigure(encoding='utf-8')


def _get_legal_rank(loai_vb: str) -> int:
    """Tính thứ bậc lập pháp chuẩn Luật Ban hành VBQPPL."""
    loai = loai_vb.lower().strip()
    if "luật" in loai or "nghị quyết quốc hội" in loai:
        return 100
    if "nghị định" in loai:
        return 200
    if "thông tư" in loai:
        return 300
    if "quy chuẩn" in loai or "tiêu chuẩn" in loai or "quyết định" in loai:
        return 400
    return 350


def get_active_legal_bases(
    dossier_type: str = "TO_TRINH_DU_TOAN",
    package_code: Optional[str] = None,
    project_context: Optional[Dict[str, Any]] = None,
    excel_path: Optional[str] = None
) -> List[Dict[str, Any]]:
    """
    Trích xuất danh sách câu căn cứ pháp lý còn hiệu lực đã được sắp xếp chuẩn thứ bậc.
    """
    resolved_excel_path = excel_path or os.path.join(
        os.path.dirname(__file__), "..", "Kho_Can_Cu_Phap_Ly.xlsx"
    )

    if not os.path.exists(resolved_excel_path):
        return []

    wb = openpyxl.load_workbook(resolved_excel_path, data_only=True)
    ws = wb.active

    # Ma trận tag cho từng loại biểu mẫu
    DOSSIER_TAG_MAP = {
        "TO_TRINH_DU_TOAN": {"DU_TOAN", "QUAN_LY_CHI_PHI", "DINH_MUC", "ALL"},
        "TO_TRINH_KHLCNT": {"DAU_THAU", "E_HSMT", "KHLCNT", "ALL"},
        "BAO_CAO_THAM_DINH": {"THAM_DINH", "QLDA", "DU_TOAN", "ALL"},
        "QD_THANH_LAP_BQLDA": {"QLDA", "BQLDA", "ALL"},
        "HOP_DONG": {"HOP_DONG", "DAU_THAU", "ALL"}
    }

    required_tags = set(DOSSIER_TAG_MAP.get(dossier_type, {"ALL"}))
    if package_code:
        required_tags.add(package_code.upper())

    context = project_context or {}
    if context.get("is_bqp_project", True):
        required_tags.add("BQP")
        required_tags.add("DOANH_TRAI")

    matched_bases = []

    for row in range(2, ws.max_row + 1):
        linh_vuc = str(ws.cell(row=row, column=2).value or "").strip()
        loai_vb = str(ws.cell(row=row, column=3).value or "").strip()
        so_hieu = str(ws.cell(row=row, column=4).value or "").strip()
        trich_yeu = str(ws.cell(row=row, column=5).value or "").strip()
        co_quan = str(ws.cell(row=row, column=6).value or "").strip()
        ngay_bh = str(ws.cell(row=row, column=7).value or "").strip()
        ngay_hl = str(ws.cell(row=row, column=8).value or "").strip()
        trang_thai = str(ws.cell(row=row, column=9).value or "").strip()
        raw_tags = str(ws.cell(row=row, column=12).value or "ALL").strip()

        if not so_hieu or not trich_yeu:
            continue

        # 1. Kiểm tra trạng thái hiệu lực (Chỉ nhận còn hiệu lực)
        if "hết hiệu lực" in trang_thai.lower() or "🔴" in trang_thai:
            continue

        # 2. Khớp Tag
        doc_tags = set([t.strip().upper() for t in raw_tags.split(",") if t.strip()])
        if "ALL" in doc_tags or doc_tags.intersection(required_tags):
            prefix = f"{loai_vb} số {so_hieu}" if loai_vb and not so_hieu.lower().startswith(loai_vb.lower()) else so_hieu
            cau_can_cu = f"Căn cứ {prefix} ngày {ngay_bh} của {co_quan} {trich_yeu};"
            thu_bac = _get_legal_rank(loai_vb)

            matched_bases.append({
                "so_hieu": so_hieu,
                "loai_vb": loai_vb,
                "trang_thai": trang_thai,
                "thu_bac": thu_bac,
                "cau_can_cu": cau_can_cu
            })

    wb.close()

    # 3. Nạp quyết định cá biệt của dự án từ context (Rank 500)
    internal_decisions = context.get("quyet_dinh_noi_bo", [])
    for qd in internal_decisions:
        matched_bases.append({
            "so_hieu": qd.get("so_hieu", "QD-NOI-BO"),
            "loai_vb": "Quyết định Chủ đầu tư",
            "trang_thai": "🟢 Còn hiệu lực",
            "thu_bac": 500,
            "cau_can_cu": f"Căn cứ Quyết định số {qd.get('so_hieu')} ngày {qd.get('ngay_bh')} của {qd.get('co_quan')} {qd.get('trich_yeu')};"
        })

    # 4. Sắp xếp theo Thứ bậc Lập pháp
    matched_bases.sort(key=lambda x: x["thu_bac"])

    # 5. Căn cứ thực tế chốt hạ (Rank 999)
    matched_bases.append({
        "so_hieu": "THUC_TE",
        "loai_vb": "Căn cứ thực tế",
        "trang_thai": "🟢 Còn hiệu lực",
        "thu_bac": 999,
        "cau_can_cu": "Căn cứ tình hình thực tế và yêu cầu triển khai thực hiện nhiệm vụ,"
    })

    return matched_bases


if __name__ == "__main__":
    bases = get_active_legal_bases(dossier_type="TO_TRINH_DU_TOAN", package_code="TV-04")
    print(f"✅ Đã trích xuất {len(bases)} căn cứ pháp lý cho gói TV-04:")
    for b in bases[:5]:
        print(f"   [{b['thu_bac']}] {b['cau_can_cu']}")
