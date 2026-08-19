#!/usr/bin/env python
# -*- coding: utf-8 -*-
"""
=============================================================================
CỖ MÁY XUẤT HỒ SƠ DỰ ÁN 1-CHẠM (DOCXTPL MASTER ENGINE)
Tác giả: Tự động hóa Hồ sơ Dự án Xây dựng
Hoạt động OFFLINE 100% trên máy tính cá nhân
=============================================================================
"""

import os
import sys
import openpyxl
from docxtpl import DocxTemplate

BASE_DIR = os.path.dirname(os.path.abspath(__file__))
EXCEL_DATA_PATH = os.path.join(BASE_DIR, "Du_lieu_mau_du_an.xlsx")
EXCEL_LEGAL_PATH = os.path.join(BASE_DIR, "Kho_Can_Cu_Phap_Ly.xlsx")
TEMPLATE_TO_TRINH = os.path.join(BASE_DIR, "Template_01_To_trinh_mau.docx")
OUTPUT_ROOT_DIR = os.path.join(BASE_DIR, "KET_QUA_XUAT_HO_SO")


def log(msg: str):
    print(f"[*] {msg}")


def load_project_data(excel_path: str) -> dict:
    wb = openpyxl.load_workbook(excel_path, data_only=True)
    context = {}

    # 1. Đọc sheet Thong_tin_chung
    if "Thong_tin_chung" in wb.sheetnames:
        ws_tt = wb["Thong_tin_chung"]
        for row in range(2, ws_tt.max_row + 1):
            var_name = ws_tt.cell(row=row, column=1).value
            var_val = ws_tt.cell(row=row, column=3).value
            if var_name and var_val is not None:
                context[str(var_name).strip()] = str(var_val).strip()

    # 2. Đọc sheet Bang_du_toan
    bang_du_toan = []
    if "Bang_du_toan" in wb.sheetnames:
        ws_dt = wb["Bang_du_toan"]
        for row in range(2, ws_dt.max_row + 1):
            stt = ws_dt.cell(row=row, column=1).value
            hang_muc = ws_dt.cell(row=row, column=2).value
            gia_tri = ws_dt.cell(row=row, column=3).value
            ghi_chu = ws_dt.cell(row=row, column=4).value
            if stt and hang_muc:
                bang_du_toan.append({
                    "stt": str(stt).strip(),
                    "hang_muc": str(hang_muc).strip(),
                    "gia_tri": str(gia_tri).strip() if gia_tri is not None else "",
                    "ghi_chu": str(ghi_chu).strip() if ghi_chu is not None else ""
                })
    context["bang_du_toan"] = bang_du_toan

    # 3. Đọc sheet Danh_muc_goi_thau
    danh_sach_goi_thau = []
    if "Danh_muc_goi_thau" in wb.sheetnames:
        ws_gt = wb["Danh_muc_goi_thau"]
        for row in range(2, ws_gt.max_row + 1):
            ma_goi = ws_gt.cell(row=row, column=1).value
            ten_goi = ws_gt.cell(row=row, column=2).value
            hinh_thuc = ws_gt.cell(row=row, column=3).value
            phuong_thuc = ws_gt.cell(row=row, column=4).value
            gia_dt = ws_gt.cell(row=row, column=5).value
            gia_tt = ws_gt.cell(row=row, column=6).value
            thoi_gian = ws_gt.cell(row=row, column=7).value
            nha_thau = ws_gt.cell(row=row, column=8).value

            if ma_goi and ten_goi:
                danh_sach_goi_thau.append({
                    "ma_goi": str(ma_goi).strip(),
                    "ten_goi": str(ten_goi).strip(),
                    "hinh_thuc": str(hinh_thuc).strip() if hinh_thuc else "",
                    "phuong_thuc": str(phuong_thuc).strip() if phuong_thuc else "",
                    "gia_du_toan": str(gia_dt).strip() if gia_dt else "",
                    "gia_trung_thau": str(gia_tt).strip() if gia_tt else "",
                    "thoi_gian": str(thoi_gian).strip() if thoi_gian else "",
                    "nha_thau": str(nha_thau).strip() if nha_thau else ""
                })
    context["danh_sach_goi_thau"] = danh_sach_goi_thau
    return context


def load_legal_bases(legal_path: str) -> dict:
    legal_context = {}
    if not os.path.exists(legal_path):
        return legal_context

    wb = openpyxl.load_workbook(legal_path, data_only=True)
    ws = wb.active
    can_cu_list = []

    for row in range(2, ws.max_row + 1):
        stt = ws.cell(row=row, column=1).value
        linh_vuc = ws.cell(row=row, column=2).value
        so_hieu = ws.cell(row=row, column=3).value
        trich_yeu = ws.cell(row=row, column=4).value
        co_quan = ws.cell(row=row, column=5).value
        ngay_bh = ws.cell(row=row, column=6).value
        trang_thai = ws.cell(row=row, column=8).value

        if so_hieu and trang_thai == "Đang có hiệu lực":
            text_can_cu = f"Căn cứ {so_hieu} ngày {ngay_bh} của {co_quan} {trich_yeu};"
            can_cu_list.append({
                "so_hieu": str(so_hieu).strip(),
                "linh_vuc": str(linh_vuc).strip() if linh_vuc else "",
                "cau_can_cu": text_can_cu
            })

    legal_context["danh_sach_can_cu_phap_ly"] = can_cu_list
    return legal_context


def generate_all_documents():
    log("=========================================================")
    log("🚀 BẮT ĐẦU QUY TRÌNH XUẤT TRỌN BỘ HỒ SƠ DỰ ÁN 1-CHẠM")
    log("=========================================================")

    if not os.path.exists(EXCEL_DATA_PATH):
        log(f"❌ Không tìm thấy file dữ liệu: {EXCEL_DATA_PATH}")
        return

    # 1. Nạp dữ liệu
    log("📊 Đang đọc dữ liệu dự án từ Excel...")
    context = load_project_data(EXCEL_DATA_PATH)
    legal_context = load_legal_bases(EXCEL_LEGAL_PATH)
    context.update(legal_context)

    log(f"   -> Dự án: {context.get('ten_du_an', 'Chưa có tên')}")
    log(f"   -> Đơn vị: {context.get('co_quan_ban_hanh', 'Chưa có')}")
    log(f"   -> Số lượng gói thầu: {len(context.get('danh_sach_goi_thau', []))}")
    log(f"   -> Số lượng mục dự toán: {len(context.get('bang_du_toan', []))}")

    # 2. Tạo thư mục kết quả
    os.makedirs(OUTPUT_ROOT_DIR, exist_ok=True)

    # 3. Xuất Tờ trình số 01 mẫu
    if os.path.exists(TEMPLATE_TO_TRINH):
        log("📄 Đang xuất Tờ trình số 01 phê duyệt nhiệm vụ và dự toán...")
        doc = DocxTemplate(TEMPLATE_TO_TRINH)
        doc.render(context)
        out_file = os.path.join(OUTPUT_ROOT_DIR, "01_To_trinh_phe_duyet_nhiem_vu_du_toan.docx")
        doc.save(out_file)
        log(f"   ✅ Đã xuất thành công: {os.path.basename(out_file)}")

    # 4. Xuất cây thư mục và danh mục hồ sơ cho từng gói thầu
    for gt in context.get("danh_sach_goi_thau", []):
        ma_goi = gt["ma_goi"]
        ten_goi = gt["ten_goi"]
        folder_name = f"Goi_{ma_goi}_{ten_goi[:30].replace(' ', '_').replace('/', '_')}"
        pkg_dir = os.path.join(OUTPUT_ROOT_DIR, folder_name)
        os.makedirs(pkg_dir, exist_ok=True)

        # Render hồ sơ riêng cho gói thầu
        pkg_context = dict(context)
        pkg_context.update({
            "ma_goi_thau": ma_goi,
            "ten_goi_thau": ten_goi,
            "hinh_thuc_lcnt": gt["hinh_thuc"],
            "gia_du_toan_goi": gt["gia_du_toan"],
            "gia_trung_thau_goi": gt["gia_trung_thau"],
            "nha_thau_goi": gt["nha_thau"],
            "thoi_gian_goi": gt["thoi_gian"]
        })

        if os.path.exists(TEMPLATE_TO_TRINH):
            doc = DocxTemplate(TEMPLATE_TO_TRINH)
            doc.render(pkg_context)
            out_file = os.path.join(pkg_dir, f"01_To_trinh_{ma_goi}.docx")
            doc.save(out_file)

        log(f"   📁 Đã cấu trúc xong thư mục gói thầu: {ma_goi} ({ten_goi[:40]}...)")

    log("=========================================================")
    log("🎉 HOÀN THÀNH XUẤT TRỌN BỘ HỒ SƠ THÀNH CÔNG 100%!")
    log(f"📂 Thư mục chứa kết quả: {OUTPUT_ROOT_DIR}")
    log("=========================================================")


if __name__ == "__main__":
    generate_all_documents()
