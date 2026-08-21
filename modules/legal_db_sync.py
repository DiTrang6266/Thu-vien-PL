# -*- coding: utf-8 -*-
"""
Module: legal_db_sync.py
Mục đích: Tự động đồng bộ CSDL Pháp lý vào file Excel (Kho_Can_Cu_Phap_Ly.xlsx).
Tính năng:
- Chuẩn hóa 14 cột dữ liệu, loại bỏ 100% tình trạng lệch cột.
- Giao diện chuyên nghiệp (Executive Styling): Freeze Panes, Auto-fit Width, Wrap Text, Badge trạng thái.
- Quản lý vòng đời (cũ hết hiệu lực / mới đang có hiệu lực).
- Safe-Write chống kẹt file trên Windows.
"""

import os
import sys
import datetime
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

if hasattr(sys.stdout, 'reconfigure'):
    sys.stdout.reconfigure(encoding='utf-8')
if hasattr(sys.stderr, 'reconfigure'):
    sys.stderr.reconfigure(encoding='utf-8')

HEADERS_14 = [
    "STT", "Lĩnh vực", "Loại văn bản", "Số hiệu", "Trích yếu", "Cơ quan ban hành",
    "Ngày ban hành", "Ngày hiệu lực", "Trạng thái", "Văn bản thay thế",
    "Quy định chuyển tiếp", "Gói thầu áp dụng", "Link Instant View", "Cập nhật lúc"
]

# Bảng màu & Định dạng chuẩn Executive Dashboard
HEADER_FILL = PatternFill(start_color="1B365D", end_color="1B365D", fill_type="solid")  # Xanh Navy Hoàng gia
HEADER_FONT = Font(name="Segoe UI", size=11, bold=True, color="FFFFFF")

ROW_EVEN_FILL = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")
ROW_ODD_FILL = PatternFill(start_color="F8FAFC", end_color="F8FAFC", fill_type="solid")

STATUS_ACTIVE_FILL = PatternFill(start_color="D1FAE5", end_color="D1FAE5", fill_type="solid")  # Xanh ngọc nhạt
STATUS_ACTIVE_FONT = Font(name="Segoe UI", size=10, bold=True, color="065F46")

STATUS_EXPIRED_FILL = PatternFill(start_color="FEE2E2", end_color="FEE2E2", fill_type="solid")  # Đỏ nhạt
STATUS_EXPIRED_FONT = Font(name="Segoe UI", size=10, bold=True, color="991B1B")

REGULAR_FONT = Font(name="Segoe UI", size=10, color="1F2937")
BOLD_FONT = Font(name="Segoe UI", size=10, bold=True, color="0F172A")
LINK_FONT = Font(name="Segoe UI", size=10, color="2563EB", underline="single")

BORDER_THIN = Border(
    left=Side(style='thin', color='CBD5E1'),
    right=Side(style='thin', color='CBD5E1'),
    top=Side(style='thin', color='CBD5E1'),
    bottom=Side(style='thin', color='CBD5E1')
)

HEADER_BORDER = Border(
    left=Side(style='thin', color='334155'),
    right=Side(style='thin', color='334155'),
    top=Side(style='thin', color='334155'),
    bottom=Side(style='medium', color='0F172A')
)

COLUMN_CONFIG = {
    1: {"width": 7, "align": Alignment(horizontal="center", vertical="center")},                          # STT
    2: {"width": 24, "align": Alignment(horizontal="left", vertical="center", wrap_text=True)},          # Lĩnh vực
    3: {"width": 14, "align": Alignment(horizontal="center", vertical="center")},                        # Loại văn bản
    4: {"width": 24, "align": Alignment(horizontal="center", vertical="center", wrap_text=True)},        # Số hiệu
    5: {"width": 46, "align": Alignment(horizontal="left", vertical="center", wrap_text=True)},          # Trích yếu
    6: {"width": 20, "align": Alignment(horizontal="center", vertical="center", wrap_text=True)},        # Cơ quan ban hành
    7: {"width": 14, "align": Alignment(horizontal="center", vertical="center")},                        # Ngày ban hành
    8: {"width": 14, "align": Alignment(horizontal="center", vertical="center")},                        # Ngày hiệu lực
    9: {"width": 18, "align": Alignment(horizontal="center", vertical="center")},                        # Trạng thái
    10: {"width": 28, "align": Alignment(horizontal="left", vertical="center", wrap_text=True)},         # Văn bản thay thế
    11: {"width": 40, "align": Alignment(horizontal="left", vertical="center", wrap_text=True)},         # Quy định chuyển tiếp
    12: {"width": 26, "align": Alignment(horizontal="left", vertical="center", wrap_text=True)},         # Gói thầu áp dụng
    13: {"width": 24, "align": Alignment(horizontal="left", vertical="center")},                          # Link Instant View
    14: {"width": 18, "align": Alignment(horizontal="center", vertical="center")}                         # Cập nhật lúc
}


def apply_table_formatting(ws) -> None:
    """Áp dụng toàn diện phong cách bảng tính chuyên nghiệp, đóng băng dòng, căn lề và màu sắc chuẩn."""
    # 1. Đóng băng dòng Tiêu đề (Freeze Panes dòng 1)
    ws.freeze_panes = "A2"

    # 2. Hiển thị đường kẻ ô rõ nét
    if hasattr(ws, "views") and ws.views.sheetView:
        ws.views.sheetView[0].showGridLines = True

    # 3. Định dạng dòng Tiêu đề (Header Row)
    ws.row_dimensions[1].height = 32
    for col_idx in range(1, 15):
        cell = ws.cell(row=1, column=col_idx)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.border = HEADER_BORDER
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    # 4. Định dạng các dòng Dữ liệu (Data Rows)
    for row_idx in range(2, ws.max_row + 1):
        ws.row_dimensions[row_idx].height = 26
        is_odd = (row_idx % 2 == 1)
        base_fill = ROW_ODD_FILL if is_odd else ROW_EVEN_FILL
        trang_thai_val = str(ws.cell(row=row_idx, column=9).value or "").strip().lower()

        for col_idx in range(1, 15):
            cell = ws.cell(row=row_idx, column=col_idx)
            cell.border = BORDER_THIN
            cfg = COLUMN_CONFIG.get(col_idx, {"align": Alignment(vertical="center")})
            cell.alignment = cfg["align"]
            cell.fill = base_fill

            # Định dạng font chữ theo từng cột
            if col_idx == 4:  # Số hiệu văn bản
                cell.font = BOLD_FONT
            elif col_idx == 1:  # STT
                cell.font = BOLD_FONT
            elif col_idx == 9:  # Trạng thái (Badge hiệu lực)
                if "hết hiệu lực" in trang_thai_val:
                    cell.fill = STATUS_EXPIRED_FILL
                    cell.font = STATUS_EXPIRED_FONT
                else:
                    cell.fill = STATUS_ACTIVE_FILL
                    cell.font = STATUS_ACTIVE_FONT
            elif col_idx == 13 and str(cell.value or "").startswith("http"):  # Link
                cell.font = LINK_FONT
            else:
                cell.font = REGULAR_FONT

    # 5. Thiết lập độ rộng cột tối ưu
    for col_idx, cfg in COLUMN_CONFIG.items():
        col_letter = get_column_letter(col_idx)
        ws.column_dimensions[col_letter].width = cfg["width"]


def clean_legal_text(raw_text: str, default: str = "") -> str:
    """Làm sạch chuỗi văn bản, loại bỏ emoji và tag thừa."""
    if not raw_text:
        return default
    text = str(raw_text).strip()
    import re
    text = re.sub(r"^[🏛📜📑🎖📐📌]\s*", "", text).strip()
    return text if text else default


def sync_legal_document_to_excel(doc_data: dict, excel_path: str = "Kho_Can_Cu_Phap_Ly.xlsx") -> bool:
    """Đồng bộ văn bản mới vào Excel, cập nhật văn bản cũ hết hiệu lực, chuẩn hóa 14 cột đẹp mắt."""
    try:
        if not os.path.exists(excel_path):
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "Can_Cu_Phap_Ly"
            ws.append(HEADERS_14)
        else:
            wb = openpyxl.load_workbook(excel_path)
            ws = wb.active

        so_hieu_moi = clean_legal_text(doc_data.get("so_hieu_clean", ""))
        vb_thay_the = clean_legal_text(doc_data.get("van_ban_thay_the", ""))
        now_str = datetime.datetime.now().strftime("%Y-%m-%d %H:%M")

        # 1. Cập nhật trạng thái văn bản cũ bị thay thế (Hỗ trợ bóc tách đa số hiệu)
        if vb_thay_the and vb_thay_the.lower() not in ["không", "chưa có", "none", ""]:
            import re
            target_numbers = re.findall(r"\d+/\d+/[A-ZĐa-zđ-]+|\d+/[A-ZĐa-zđ-]+|TCVN\s*\d+(?::\d+)?|QCVN\s*\d+(?::\d+)?", vb_thay_the)
            target_numbers_clean = [n.strip().lower().replace(" ", "") for n in target_numbers]

            for r in range(2, ws.max_row + 1):
                cell_so_hieu = str(ws.cell(row=r, column=4).value or "").strip()
                cell_so_hieu_clean = cell_so_hieu.lower().replace(" ", "")

                is_matched = False
                if cell_so_hieu_clean:
                    if cell_so_hieu_clean in vb_thay_the.lower() or any(tn in cell_so_hieu_clean or cell_so_hieu_clean in tn for tn in target_numbers_clean):
                        is_matched = True

                if is_matched:
                    ws.cell(row=r, column=9).value = "Hết hiệu lực"
                    current_repl = str(ws.cell(row=r, column=10).value or "").strip()
                    new_repl_note = f"Bị thay thế bởi {so_hieu_moi}"
                    if current_repl and current_repl.lower() not in ["none", ""] and new_repl_note.lower() not in current_repl.lower():
                        ws.cell(row=r, column=10).value = f"{current_repl}, {new_repl_note}"
                    else:
                        ws.cell(row=r, column=10).value = new_repl_note

        # 2. Kiểm tra nếu văn bản mới đã tồn tại thì bỏ qua (chống trùng lặp)
        for r in range(2, ws.max_row + 1):
            if str(ws.cell(row=r, column=4).value or "").strip().lower() == so_hieu_moi.lower():
                apply_table_formatting(ws)
                wb.save(excel_path)
                wb.close()
                return True

        # 3. Thêm dòng văn bản mới chuẩn 14 cột
        next_stt = ws.max_row
        tags_raw = doc_data.get("goi_thau_tags", ["ALL"])
        tags_str = ", ".join(tags_raw) if isinstance(tags_raw, list) else str(tags_raw)

        linh_vuc = clean_legal_text(doc_data.get("linh_vuc", "Xây dựng & Đấu thầu"))
        loai_vb = clean_legal_text(doc_data.get("loai_van_ban", "Văn bản QPPL"))
        co_quan = clean_legal_text(doc_data.get("co_quan", "Nhà nước"))
        trich_yeu = clean_legal_text(doc_data.get("trich_yeu", doc_data.get("title", "")))

        new_row = [
            next_stt,
            linh_vuc,
            loai_vb,
            so_hieu_moi,
            trich_yeu,
            co_quan,
            doc_data.get("ngay_ban_hanh", ""),
            doc_data.get("ngay_hieu_luc", ""),
            "Đang có hiệu lực",
            vb_thay_the,
            doc_data.get("chuyen_tiep_ngan", ""),
            tags_str,
            doc_data.get("telegraph_url", ""),
            now_str
        ]
        ws.append(new_row)

        # 4. Đánh lại STT từ 1 đến N
        for r in range(2, ws.max_row + 1):
            ws.cell(row=r, column=1).value = r - 1

        # 5. Áp dụng phong cách bảng tính chuyên nghiệp
        apply_table_formatting(ws)

        # 6. Lưu an toàn (Safe-Write chống kẹt file khi đang mở Excel)
        try:
            wb.save(excel_path)
        except PermissionError:
            pending_path = excel_path.replace(".xlsx", "_pending.xlsx")
            wb.save(pending_path)
            print(f"[!] CẢNH BÁO: Excel đang mở. Đã lưu tạm an toàn vào: {os.path.basename(pending_path)}")

        wb.close()

        # 7. Tự động cập nhật Trang Web Thẻ Di Động (GitHub Pages)
        try:
            from modules.web_card_generator import generate_mobile_card_web
            generate_mobile_card_web(excel_path=excel_path)
        except Exception as e:
            print(f"[!] Không thể sinh Web Card View: {e}")

        return True
    except Exception as e:
        print(f"[X] Lỗi đồng bộ Excel: {e}")
        return False
