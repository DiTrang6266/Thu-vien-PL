# -*- coding: utf-8 -*-
"""
Module: web_card_generator.py
Mục đích: Tự động chuyển đổi Sổ cái Excel (Kho_Can_Cu_Phap_Ly.xlsx) thành Trang Web Thẻ Di Động Thông Minh (Mobile Card-View).
Tính năng:
1. Giao diện Thẻ Dọc (Vertical Cards) tối ưu 100% cho màn hình điện thoại di động (không bị tràn ngang).
2. Ô tìm kiếm tức thì (Live Search) theo số hiệu, trích yếu, cơ quan ban hành.
3. Thanh lọc 1-chạm (Filter Pills): Tất cả, Xây lắp, Tư vấn, Doanh cụ, Chỉ xem văn bản Còn hiệu lực.
4. Nút bấm 1-Chạm Sao Chép (1-Tap Copy) câu căn cứ chuẩn Nghị định 30 vào Clipboard kèm Toast thông báo.
5. Chi phí 0đ vĩnh viễn, lưu trữ trên GitHub Pages (https://ditrang6266.github.io/Thu-vien-PL/).
"""

import os
import sys
import json
import openpyxl

if hasattr(sys.stdout, 'reconfigure'):
    sys.stdout.reconfigure(encoding='utf-8')


HTML_TEMPLATE = """<!DOCTYPE html>
<html lang="vi">
<head>
    <meta charset="UTF-8">
    <meta name="viewport" content="width=device-width, initial-scale=1.0, maximum-scale=1.0, user-scalable=no">
    <title>🏛️ Thư Viện Căn Cứ Pháp Lý Dự Án</title>
    <style>
        :root {
            --primary: #1a56db;
            --primary-light: #e1effe;
            --success-bg: #def7ec;
            --success-text: #03543f;
            --danger-bg: #fde8e8;
            --danger-text: #9b1c1c;
            --card-bg: #ffffff;
            --bg: #f3f4f6;
            --text-main: #111827;
            --text-muted: #6b7280;
            --border: #e5e7eb;
        }
        * { box-sizing: border-box; margin: 0; padding: 0; font-family: -apple-system, BlinkMacSystemFont, "Segoe UI", Roboto, Helvetica, Arial, sans-serif; }
        body { background: var(--bg); color: var(--text-main); padding-bottom: 80px; -webkit-font-smoothing: antialiased; }
        
        /* Header */
        .header { background: #1e293b; color: #fff; padding: 18px 16px; position: sticky; top: 0; z-index: 100; box-shadow: 0 2px 8px rgba(0,0,0,0.15); }
        .header-title { font-size: 1.15rem; font-weight: 700; display: flex; align-items: center; justify-content: space-between; }
        .badge-live { background: #10b981; color: #fff; font-size: 0.7rem; padding: 2px 8px; border-radius: 99px; font-weight: 600; }
        .header-sub { font-size: 0.8rem; color: #94a3b8; margin-top: 4px; }
        
        /* Search & Filter Container */
        .sticky-tools { background: #ffffff; padding: 12px 16px; border-bottom: 1px solid var(--border); position: sticky; top: 0; z-index: 90; box-shadow: 0 2px 6px rgba(0,0,0,0.04); }
        .search-box { width: 100%; padding: 10px 14px 10px 36px; border: 1.5px solid #cbd5e1; border-radius: 10px; font-size: 0.95rem; outline: none; background: #f8fafc url('data:image/svg+xml;utf8,<svg xmlns="http://www.w3.org/2000/svg" width="16" height="16" fill="%2364748b" viewBox="0 0 16 16"><path d="M11.742 10.344a6.5 6.5 0 1 0-1.397 1.398h-.001c.03.04.062.078.098.115l3.85 3.85a1 1 0 0 0 1.415-1.414l-3.85-3.85a1.007 1.007 0 0 0-.115-.1zM12 6.5a5.5 5.5 0 1 1-11 0 5.5 5.5 0 0 1 11 0z"/></svg>') no-repeat 12px center; transition: all 0.2s; }
        .search-box:focus { border-color: var(--primary); background-color: #fff; box-shadow: 0 0 0 3px rgba(26,86,219,0.15); }
        
        /* Filter Pills */
        .filter-section { display: flex; flex-direction: column; gap: 8px; margin-top: 10px; }
        .filter-group { display: flex; flex-wrap: wrap; gap: 6px; }
        
        /* Slider with Navigation Arrows < > */
        .nav-slider-wrapper { display: flex; align-items: center; gap: 6px; position: relative; }
        .nav-arrow-btn {
            background: #ffffff;
            color: #1e293b;
            border: 1.5px solid #cbd5e1;
            border-radius: 8px;
            width: 32px;
            height: 32px;
            min-width: 32px;
            font-size: 1.25rem;
            font-weight: 700;
            line-height: 1;
            display: flex;
            align-items: center;
            justify-content: center;
            cursor: pointer;
            user-select: none;
            transition: all 0.15s ease;
            box-shadow: 0 1px 3px rgba(0,0,0,0.06);
        }
        .nav-arrow-btn:hover {
            background: var(--primary);
            color: #ffffff;
            border-color: var(--primary);
            box-shadow: 0 2px 5px rgba(26,86,219,0.25);
        }
        .nav-arrow-btn:active { transform: scale(0.92); }
        
        .pills-scroll-track {
            display: flex;
            gap: 6px;
            overflow-x: auto;
            scroll-behavior: smooth;
            padding: 2px 2px;
            scrollbar-width: none;
            -webkit-overflow-scrolling: touch;
            cursor: grab;
            flex: 1;
        }
        .pills-scroll-track:active { cursor: grabbing; }
        .pills-scroll-track::-webkit-scrollbar { display: none; }

        .pill { user-select: none; -webkit-user-select: none; white-space: nowrap; padding: 6px 12px; border-radius: 8px; font-size: 0.8rem; font-weight: 600; border: 1px solid #cbd5e1; background: #f8fafc; color: #334155; cursor: pointer; transition: all 0.15s; flex-shrink: 0; }
        .pill:hover { background: #f1f5f9; border-color: #94a3b8; }
        .pill.active { background: var(--primary); color: #fff; border-color: var(--primary); box-shadow: 0 2px 4px rgba(26,86,219,0.25); }
        .pill-sub { font-size: 0.78rem; padding: 5px 11px; }

        /* Stats bar */
        .stats-bar { padding: 10px 16px 4px 16px; font-size: 0.8rem; color: var(--text-muted); font-weight: 500; }

        /* Card List */
        .card-list { padding: 8px 16px; display: flex; flex-direction: column; gap: 14px; }
        .legal-card { background: var(--card-bg); border-radius: 14px; padding: 16px; border: 1px solid var(--border); box-shadow: 0 1px 3px rgba(0,0,0,0.05); transition: transform 0.1s, box-shadow 0.1s; }
        .card-header { display: flex; justify-content: space-between; align-items: flex-start; gap: 8px; margin-bottom: 8px; }
        .so-hieu { font-size: 1.05rem; font-weight: 700; color: #0f172a; word-break: break-word; }
        .status-badge { font-size: 0.75rem; font-weight: 700; padding: 3px 8px; border-radius: 6px; white-space: nowrap; }
        .status-active { background: var(--success-bg); color: var(--success-text); }
        .status-expired { background: var(--danger-bg); color: var(--danger-text); }

        .meta-row { display: flex; flex-wrap: wrap; gap: 10px; font-size: 0.8rem; color: #475569; margin-bottom: 10px; }
        .meta-item { display: flex; align-items: center; gap: 4px; }

        .trich-yeu { font-size: 0.9rem; color: #1e293b; line-height: 1.45; margin-bottom: 12px; font-weight: 500; }
        
        .info-box { background: #f8fafc; border-left: 3.5px solid #94a3b8; padding: 8px 10px; border-radius: 4px; font-size: 0.8rem; color: #334155; margin-bottom: 12px; line-height: 1.4; }
        .info-box.transition { border-left-color: #f59e0b; background: #fffbeb; }

        .tags-row { display: flex; flex-wrap: wrap; gap: 6px; margin-bottom: 14px; }
        .tag-pill { background: #e0f2fe; color: #0369a1; font-size: 0.75rem; font-weight: 600; padding: 2px 8px; border-radius: 4px; }

        /* Action Buttons */
        .card-actions { display: grid; grid-template-columns: 1fr auto; gap: 8px; }
        .btn-copy { background: var(--primary-light); color: var(--primary); border: 1px solid rgba(26,86,219,0.3); font-weight: 600; font-size: 0.85rem; padding: 9px 12px; border-radius: 8px; cursor: pointer; display: flex; align-items: center; justify-content: center; gap: 6px; transition: all 0.15s; }
        .btn-copy:active { background: var(--primary); color: #fff; }
        .btn-view { background: #f1f5f9; color: #334155; border: 1px solid #cbd5e1; font-weight: 600; font-size: 0.85rem; padding: 9px 12px; border-radius: 8px; text-decoration: none; display: flex; align-items: center; justify-content: center; gap: 4px; }

        /* Toast */
        .toast { position: fixed; bottom: 24px; left: 50%; transform: translateX(-50%) translateY(100px); background: #0f172a; color: #fff; padding: 12px 20px; border-radius: 99px; font-size: 0.85rem; font-weight: 600; box-shadow: 0 10px 25px rgba(0,0,0,0.3); z-index: 999; opacity: 0; transition: all 0.3s cubic-bezier(0.68, -0.55, 0.265, 1.55); pointer-events: none; text-align: center; }
        .toast.show { transform: translateX(-50%) translateY(0); opacity: 1; }

        .empty-state { text-align: center; padding: 40px 20px; color: var(--text-muted); font-size: 0.95rem; }
    </style>
</head>
<body>

    <!-- Header -->
    <div class="header">
        <div class="header-title">
            <span>🏛️ KHO CĂN CỨ PHÁP LÝ</span>
            <span class="badge-live">TỰ ĐỘNG 24/7</span>
        </div>
        <div class="header-sub">Hồ sơ Dự án Xây dựng & Đấu thầu</div>
    </div>

    <!-- Search & Filter Controls -->
    <div class="sticky-tools">
        <input type="text" id="searchInput" class="search-box" placeholder="Tìm số hiệu (24/2024...), định mức, tên văn bản..." oninput="renderFilteredCards()">
        <div class="filter-section">
            <div class="filter-group">
                <button class="pill active" onclick="setFilterCategory('ALL', this)">Tất cả (94)</button>
                <button class="pill" onclick="setFilterCategory('HIEU_LUC', this)">🟢 Còn hiệu lực</button>
                <button class="pill" onclick="setFilterCategory('HET_HIEU_LUC', this)">🔴 Hết hiệu lực</button>
            </div>
            <div class="nav-slider-wrapper">
                <button class="nav-arrow-btn" onclick="scrollPills(-240)" title="Cuộn sang trái">‹</button>
                <div class="pills-scroll-track" id="pillsTrack">
                    <button class="pill pill-sub" onclick="setFilterCategory('QUY_HOACH_KHAO_SAT', this)">🗺️ Quy hoạch & Khảo sát</button>
                    <button class="pill pill-sub" onclick="setFilterCategory('THIET_KE_DU_TOAN', this)">📐 Thiết kế & Dự toán</button>
                    <button class="pill pill-sub" onclick="setFilterCategory('THAM_TRA_THAM_DINH', this)">🔍 Thẩm tra & Thẩm định</button>
                    <button class="pill pill-sub" onclick="setFilterCategory('DAU_THAU_HOP_DONG', this)">⚖️ Đấu thầu & Hợp đồng</button>
                    <button class="pill pill-sub" onclick="setFilterCategory('GIAM_SAT_CHAT_LUONG', this)">👷 Giám sát & Nghiệm thu</button>
                    <button class="pill pill-sub" onclick="setFilterCategory('THI_CONG_XAY_DUNG', this)">🏗️ Thi công & An toàn</button>
                    <button class="pill pill-sub" onclick="setFilterCategory('KIEM_TOAN_QUYET_TOAN', this)">📊 Kiểm toán & Quyết toán</button>
                    <button class="pill pill-sub" onclick="setFilterCategory('BAO_HIEM', this)">🛡️ Bảo hiểm công trình</button>
                    <button class="pill pill-sub" onclick="setFilterCategory('BQP', this)">🎖️ Bộ Quốc phòng</button>
                    <button class="pill pill-sub" onclick="setFilterCategory('PCCC', this)">🔥 PCCC</button>
                    <button class="pill pill-sub" onclick="setFilterCategory('MOI_TRUONG', this)">🌿 Môi trường</button>
                    <button class="pill pill-sub" onclick="setFilterCategory('CHI_THUONG_XUYEN', this)">💼 Chi thường xuyên</button>
                </div>
                <button class="nav-arrow-btn" onclick="scrollPills(240)" title="Cuộn sang phải">›</button>
            </div>
        </div>
    </div>

    <!-- Stats -->
    <div class="stats-bar" id="statsBar">Đang tải dữ liệu...</div>

    <!-- Card List Container -->
    <div class="card-list" id="cardList"></div>

    <!-- Toast Notification -->
    <div class="toast" id="toast">✅ Đã sao chép câu căn cứ vào Clipboard!</div>

    <script>
        const RAW_DATA = ###DATA_PLACEHOLDER###;
        let currentFilter = 'ALL';

        function scrollPills(distance) {
            const track = document.getElementById('pillsTrack');
            if (track) {
                track.scrollBy({ left: distance, behavior: 'smooth' });
            }
        }

        function setFilterCategory(cat, el) {
            currentFilter = cat;
            document.querySelectorAll('.pill').forEach(p => p.classList.remove('active'));
            if (el) {
                el.classList.add('active');
                if (el.parentElement && el.parentElement.id === 'pillsTrack') {
                    el.scrollIntoView({ behavior: 'smooth', inline: 'center', block: 'nearest' });
                }
            } else {
                const btn = Array.from(document.querySelectorAll('.pill')).find(b => b.getAttribute('onclick') && b.getAttribute('onclick').includes(`'${cat}'`));
                if (btn) {
                    btn.classList.add('active');
                    if (btn.parentElement && btn.parentElement.id === 'pillsTrack') {
                        btn.scrollIntoView({ behavior: 'smooth', inline: 'center', block: 'nearest' });
                    }
                }
            }
            renderFilteredCards();
        }

        function filterByTag(tag) {
            tag = tag.trim().replace(/^#/, '');
            setFilterCategory(tag, null);
            window.scrollTo({ top: 0, behavior: 'smooth' });
        }

        // Kích hoạt tính năng lăn chuột và kéo rê chuột cho máy tính
        document.addEventListener('DOMContentLoaded', () => {
            const track = document.getElementById('pillsTrack');
            if (!track) return;

            // 1. Lăn con lăn chuột để cuộn ngang
            track.addEventListener('wheel', (e) => {
                if (e.deltaY !== 0) {
                    e.preventDefault();
                    track.scrollLeft += e.deltaY * 1.2;
                }
            }, { passive: false });

            // 2. Kéo rê chuột (Drag-to-Scroll)
            let isDown = false;
            let startX;
            let scrollLeft;

            track.addEventListener('mousedown', (e) => {
                isDown = true;
                track.classList.add('active');
                startX = e.pageX - track.offsetLeft;
                scrollLeft = track.scrollLeft;
            });
            track.addEventListener('mouseleave', () => { isDown = false; });
            track.addEventListener('mouseup', () => { isDown = false; });
            track.addEventListener('mousemove', (e) => {
                if (!isDown) return;
                e.preventDefault();
                const x = e.pageX - track.offsetLeft;
                const walk = (x - startX) * 1.5;
                track.scrollLeft = scrollLeft - walk;
            });
        });

        function renderFilteredCards() {
            const query = (document.getElementById('searchInput').value || '').toLowerCase().trim();
            const listEl = document.getElementById('cardList');
            const statsEl = document.getElementById('statsBar');

            const filtered = RAW_DATA.filter(item => {
                // 1. Text search
                const matchText = (
                    item.so_hieu.toLowerCase().includes(query) ||
                    item.trich_yeu.toLowerCase().includes(query) ||
                    item.co_quan.toLowerCase().includes(query) ||
                    (item.linh_vuc || '').toLowerCase().includes(query) ||
                    (item.tags || '').toLowerCase().includes(query)
                );
                if (!matchText) return false;

                // 2. Category filter
                if (currentFilter === 'ALL') return true;
                if (currentFilter === 'HIEU_LUC') return item.trang_thai.toLowerCase().includes('đang có hiệu lực');
                if (currentFilter === 'HET_HIEU_LUC') return !item.trang_thai.toLowerCase().includes('đang có hiệu lực');
                
                const itemTags = (item.tags || '').toUpperCase();
                const linhVuc = (item.linh_vuc || '').toLowerCase();
                const trichYeu = (item.trich_yeu || '').toLowerCase();
                const coQuan = (item.co_quan || '').toLowerCase();
                const filterUpper = currentFilter.toUpperCase();

                if (filterUpper === 'QUY_HOACH_KHAO_SAT') {
                    return linhVuc.includes('quy hoạch') || linhVuc.includes('khảo sát') || linhVuc.includes('đo đạc') ||
                           itemTags.includes('QUY_HOACH') || itemTags.includes('KHAO_SAT') || itemTags.includes('DO_DAC') || itemTags.includes('TV-01') || itemTags.includes('TV-02') ||
                           trichYeu.includes('quy hoạch') || trichYeu.includes('khảo sát') || trichYeu.includes('đo đạc') || trichYeu.includes('địa chất') || trichYeu.includes('trắc địa');
                }
                if (filterUpper === 'THIET_KE_DU_TOAN') {
                    return linhVuc.includes('chi phí') || linhVuc.includes('định mức') || linhVuc.includes('thiết kế') ||
                           itemTags.includes('CHI_PHI') || itemTags.includes('DINH_MUC') || itemTags.includes('THIET_KE') || itemTags.includes('DU_TOAN') || itemTags.includes('TV-04') ||
                           trichYeu.includes('thiết kế') || trichYeu.includes('dự toán') || trichYeu.includes('định mức') || trichYeu.includes('quản lý chi phí');
                }
                if (filterUpper === 'THAM_TRA_THAM_DINH') {
                    return itemTags.includes('THAM_TRA') || itemTags.includes('THAM_DINH') || itemTags.includes('TV-03') || itemTags.includes('TV-05') ||
                           trichYeu.includes('thẩm tra') || trichYeu.includes('thẩm định') || trichYeu.includes('chủ trương') || linhVuc.includes('thẩm định');
                }
                if (filterUpper === 'DAU_THAU_HOP_DONG') {
                    return linhVuc.includes('đấu thầu') || linhVuc.includes('hợp đồng') ||
                           itemTags.includes('DAU_THAU') || itemTags.includes('HOP_DONG') || itemTags.includes('TV-06') ||
                           trichYeu.includes('đấu thầu') || trichYeu.includes('hợp đồng') || trichYeu.includes('lựa chọn nhà thầu') || trichYeu.includes('tạm ứng');
                }
                if (filterUpper === 'GIAM_SAT_CHAT_LUONG') {
                    return linhVuc.includes('chất lượng') || linhVuc.includes('nghiệm thu') || linhVuc.includes('giám sát') ||
                           itemTags.includes('GIAM_SAT') || itemTags.includes('CHAT_LUONG') || itemTags.includes('NGHIEM_THU') || itemTags.includes('TV-08') ||
                           trichYeu.includes('giám sát') || trichYeu.includes('chất lượng') || trichYeu.includes('nghiệm thu') || trichYeu.includes('nhật ký');
                }
                if (filterUpper === 'THI_CONG_XAY_DUNG') {
                    return linhVuc.includes('xây dựng') || linhVuc.includes('an toàn') || linhVuc.includes('tiêu chuẩn') ||
                           itemTags.includes('THI_CONG') || itemTags.includes('AN_TOAN') || itemTags.includes('XD-01') || itemTags.includes('TV-07') || itemTags.includes('THI_NGHIEM_COC') ||
                           trichYeu.includes('thi công') || trichYeu.includes('an toàn lao động') || trichYeu.includes('nén cọc') || trichYeu.includes('thí nghiệm');
                }
                if (filterUpper === 'KIEM_TOAN_QUYET_TOAN') {
                    return linhVuc.includes('kiểm toán') || linhVuc.includes('quyết toán') ||
                           itemTags.includes('KIEM_TOAN') || itemTags.includes('QUYET_TOAN') || itemTags.includes('TV-09') ||
                           trichYeu.includes('quyết toán') || trichYeu.includes('kiểm toán');
                }
                if (filterUpper === 'BAO_HIEM') {
                    return linhVuc.includes('bảo hiểm') || itemTags.includes('BAO_HIEM') || itemTags.includes('PTV-01') || trichYeu.includes('bảo hiểm');
                }
                if (filterUpper === 'BQP') {
                    return itemTags.includes('BQP') || coQuan.includes('quốc phòng') || linhVuc.includes('quốc phòng') || linhVuc.includes('doanh trại') || trichYeu.includes('bộ quốc phòng') || trichYeu.includes('doanh trại');
                }
                if (filterUpper === 'PCCC') {
                    return itemTags.includes('PCCC') || linhVuc.includes('pccc') || linhVuc.includes('phòng cháy') || trichYeu.includes('cháy') || trichYeu.includes('pccc') || trichYeu.includes('cứu nạn');
                }
                if (filterUpper === 'MOI_TRUONG') {
                    return itemTags.includes('MOI_TRUONG') || linhVuc.includes('môi trường') || trichYeu.includes('môi trường') || trichYeu.includes('đtm');
                }
                if (filterUpper === 'CHI_THUONG_XUYEN') {
                    return itemTags.includes('CHI_THUONG_XUYEN') || linhVuc.includes('chi thường xuyên') || linhVuc.includes('tài sản') || trichYeu.includes('chi thường xuyên') || trichYeu.includes('tài sản công') || trichYeu.includes('sửa chữa');
                }

                return itemTags.includes(filterUpper);
            });

            statsEl.innerText = `Hiển thị ${filtered.length} / ${RAW_DATA.length} văn bản pháp lý`;

            if (filtered.length === 0) {
                listEl.innerHTML = '<div class="empty-state">🔍 Không tìm thấy văn bản pháp luật nào phù hợp.</div>';
                return;
            }

            listEl.innerHTML = filtered.map(item => {
                const isActive = item.trang_thai.toLowerCase().includes('đang có hiệu lực');
                const statusClass = isActive ? 'status-active' : 'status-expired';
                const statusText = isActive ? '🟢 CÒN HIỆU LỰC' : '🔴 HẾT HIỆU LỰC';

                const tagsList = (item.tags || 'ALL').split(',')
                    .map(t => t.trim())
                    .filter(t => t && !t.startsWith('TV-') && !t.startsWith('PTV-') && !t.startsWith('XD-'))
                    .map(cleanTag => {
                        return `<span class="tag-pill" style="cursor:pointer;" onclick="filterByTag('${cleanTag}')">${cleanTag}</span>`;
                    }).join('');

                let transitionHtml = '';
                if (item.chuyen_tiep && item.chuyen_tiep.length > 5) {
                    transitionHtml = `<div class="info-box transition">⚡ <b>Chuyển tiếp:</b> ${item.chuyen_tiep}</div>`;
                } else if (item.thay_the && item.thay_the.length > 3) {
                    transitionHtml = `<div class="info-box">🔄 <b>Thay thế:</b> ${item.thay_the}</div>`;
                }

                const instantViewBtn = item.link_iv ? `<a href="${item.link_iv}" target="_blank" class="btn-view">⚡ Đọc Báo Cáo</a>` : '';

                // Escape double quotes for JS copy string
                const safeClause = (item.cau_can_cu || `Căn cứ ${item.so_hieu} ngày ${item.ngay_bh} của ${item.co_quan} ${item.trich_yeu};`).replace(/"/g, '&quot;');

                return `
                <div class="legal-card">
                    <div class="card-header">
                        <div class="so-hieu">${item.so_hieu}</div>
                        <span class="status-badge ${statusClass}">${statusText}</span>
                    </div>
                    <div class="meta-row">
                        <span class="meta-item">🏢 ${item.co_quan}</span>
                        <span class="meta-item">📅 Ban hành: ${item.ngay_bh}</span>
                        ${item.ngay_hl ? `<span class="meta-item">⏱ Hiệu lực: ${item.ngay_hl}</span>` : ''}
                    </div>
                    <div class="trich-yeu">${item.trich_yeu}</div>
                    ${transitionHtml}
                    <div class="tags-row">${tagsList}</div>
                    <div class="card-actions">
                        <button class="btn-copy" onclick="copyClause('${safeClause}')">
                            📋 Sao Chép Căn Cứ
                        </button>
                        ${instantViewBtn}
                    </div>
                </div>
                `;
            }).join('');
        }

        function copyClause(text) {
            navigator.clipboard.writeText(text).then(() => {
                showToast('✅ Đã sao chép câu căn cứ vào Clipboard!');
            }).catch(() => {
                // Fallback
                const t = document.createElement("textarea");
                t.value = text;
                document.body.appendChild(t);
                t.select();
                document.execCommand('copy');
                document.body.removeChild(t);
                showToast('✅ Đã sao chép câu căn cứ vào Clipboard!');
            });
        }

        function showToast(msg) {
            const toast = document.getElementById('toast');
            toast.innerText = msg;
            toast.classList.add('show');
            setTimeout(() => { toast.classList.remove('show'); }, 2200);
        }

        // Init on load
        renderFilteredCards();
    </script>
</body>
</html>
"""


def generate_mobile_card_web(
    excel_path: str = "Kho_Can_Cu_Phap_Ly.xlsx",
    output_dir: str = "docs"
) -> str:
    """Đọc file Excel và sinh ra file index.html giao diện Thẻ Di Động Thông Minh."""
    if not os.path.exists(excel_path):
        return ""

    wb = openpyxl.load_workbook(excel_path, data_only=True)
    ws = wb.active
    records = []

    for r in range(2, ws.max_row + 1):
        so_hieu = str(ws.cell(row=r, column=4).value or "").strip()
        if not so_hieu:
            continue

        linh_vuc = str(ws.cell(row=r, column=2).value or "").strip()
        loai_vb = str(ws.cell(row=r, column=3).value or "").strip()
        trich_yeu = str(ws.cell(row=r, column=5).value or "").strip()
        co_quan = str(ws.cell(row=r, column=6).value or "").strip()
        ngay_bh = str(ws.cell(row=r, column=7).value or "").strip()
        ngay_hl = str(ws.cell(row=r, column=8).value or "").strip()
        trang_thai = str(ws.cell(row=r, column=9).value or "Đang có hiệu lực").strip()
        thay_the = str(ws.cell(row=r, column=10).value or "").strip()
        chuyen_tiep = str(ws.cell(row=r, column=11).value or "").strip()
        tags = str(ws.cell(row=r, column=12).value or "ALL").strip()
        link_iv = str(ws.cell(row=r, column=13).value or "").strip()

        # Tạo câu căn cứ chuẩn Nghị định 30
        prefix = f"{loai_vb} số {so_hieu}" if loai_vb and not so_hieu.lower().startswith(loai_vb.lower()) else so_hieu
        cau_can_cu = f"Căn cứ {prefix} ngày {ngay_bh} của {co_quan} {trich_yeu};"

        records.append({
            "stt": r - 1,
            "linh_vuc": linh_vuc,
            "loai_vb": loai_vb,
            "so_hieu": so_hieu,
            "trich_yeu": trich_yeu,
            "co_quan": co_quan,
            "ngay_bh": ngay_bh,
            "ngay_hl": ngay_hl,
            "trang_thai": trang_thai,
            "thay_the": thay_the,
            "chuyen_tiep": chuyen_tiep,
            "tags": tags,
            "link_iv": link_iv,
            "cau_can_cu": cau_can_cu
        })

    wb.close()

    # Đảo ngược để các văn bản mới nhất nằm ở trên đầu
    records.reverse()

    json_payload = json.dumps(records, ensure_ascii=False)
    final_html = HTML_TEMPLATE.replace("###DATA_PLACEHOLDER###", json_payload)

    # 1. Ghi vào thư mục docs/index.html (Chuẩn GitHub Pages)
    os.makedirs(output_dir, exist_ok=True)
    docs_html_path = os.path.join(output_dir, "index.html")
    with open(docs_html_path, "w", encoding="utf-8") as f:
        f.write(final_html)

    # 2. Ghi song song vào thư mục gốc index.html
    root_html_path = "index.html"
    with open(root_html_path, "w", encoding="utf-8") as f:
        f.write(final_html)

    return docs_html_path


if __name__ == "__main__":
    out = generate_mobile_card_web()
    print(f"✅ Đã sinh Trang Web Thẻ Di Động thành công: {out}")
